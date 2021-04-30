from PyQt5.QtCore import Qt, QThread, pyqtSignal

from openpyxl import Workbook
from datetime import datetime, timedelta

class SerialThread(QThread):
    signal = pyqtSignal(str)

    def __init__(self, parent, arduino_connection):
        QThread.__init__(self)


        # stuff
        self.arduino_started_timestamp = 0
        self.workbook = Workbook()

        self.pre_trial_sheet = self.workbook.active
        self.pre_trial_sheet.title = "Pre-trial"
 
        #self.pre_trial_sheet = self.workbook.create_sheet("Pre-trial")
        
        #headers = ["Code", "Timestamp", "Delta start timestamp"]
        #self.pre_trial_sheet.append(headers)

        self.trial = 0
        
        # start Arduino and record data and stop recording data
        self.start_signal = False
        self.end_signal = False
        
        self.arduino_connection = arduino_connection
        self.buffer = ''
        self.expected_strings = ['HL_ON','HL_OFF', 
                                 'LL_ON', 'LL_OFF', 
                                 'RL_ON', 'RL_OFF',
                                 'PR', 'OR', 'TR', 'SR',
                                 'ITIS', 'DS', 'MO',
                                 'FSA', 'FSR', 'FSF', 'LSA',
                                 'LSR', 'RSA', 'RSR', 
                                 'SA', 'TO', 'RTS', 'FTS', 'TE', 'ART']

    def run(self):
        while True:
            if self.start_signal:
                while self.arduino_connection.readline().decode("utf-8").strip() != "SA":
                    print('Trying to start Arduino')
                    self.arduino_connection.write(bytes("start", 'utf-8'))
                    self.start_signal = False
                
                self.arduino_started_timestamp = datetime.now()
                self.write_to_spreadsheet(self.trial, "SA", self.arduino_started_timestamp)
                self.signal.emit("SA")
            else:
                if not self.end_signal:
                    decodedStream = self.arduino_connection.readline().decode("utf-8").strip()
                    if decodedStream != '':
                        # need to create a buffer to avoid getting just partial strings
                        self.buffer = self.buffer + decodedStream

                        if self.buffer in self.expected_strings:
                            self.write_to_spreadsheet(self.trial, self.buffer, datetime.now())
                            if self.buffer == 'TE' or self.buffer == "ART":
                                self.trial = self.trial + 1
                                new_sheet_name = "Trial " + str(self.trial)
                                self.workbook.create_sheet(new_sheet_name)

                            self.signal.emit(self.buffer)
                            self.buffer = ""

    def start_trial(self):
        self.start_signal = True

    def end_trial(self):
        self.end_signal = True
        self.workbook.save(filename=self.arduino_started_timestamp.strftime("%a %d %b %Y %H %M %S") + ".xlsx")

    def write_to_spreadsheet(self, sheet_number, code, timestamp):
        #print(self.workbook.get_sheet_names())
        #print(sheet_number)
        #print(self.workbook.get_sheet_names()[sheet_number])
        #print([code, timestamp, timestamp - self.arduino_started_timestamp])
        
        working_sheet = self.workbook.get_sheet_by_name(self.workbook.get_sheet_names()[sheet_number])
        working_sheet.append([code, timestamp.strftime("%H:%M:%S:%f"), str(timestamp - self.arduino_started_timestamp)])

        #self.spreadsheet.save(filename="olalalal.xlsx")