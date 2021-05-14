"""
creation date: 5-4-2021 20:00
last modified: 07-05-2021 16:15

@author: Bruno Correia - brunorisico@gmail.com

Thread used to connect to the Arduino and to generate the sheets
"""

from PyQt5.QtCore import Qt, QThread, pyqtSignal

from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from random import randrange

class SerialThread(QThread):
    signal = pyqtSignal(str)

    def __init__(self, parent, arduino_connection, vds_name):
        QThread.__init__(self)

        self.vds_name = vds_name
        # stuff
        self.arduino_started_timestamp = 0
        self.workbook = Workbook()

        self.pre_trial_sheet = self.workbook.active
        self.pre_trial_sheet.title = "Pre-trial"
 
        self.headers = ["Code", "Timestamp", "Delta start timestamp"]
        self.pre_trial_sheet.append(self.headers)

        self.trial = 0
        self.left_sensor_activations = 0
        self.right_sensor_activations = 0
        self.feeder_sensor_activations = 0

        # start Arduino and record data and stop recording data
        self.start_signal = False
        self.end_signal = False
        
        self.arduino_connection = arduino_connection
        self.buffer = ''
        self.expected_strings = ['HL_ON','HL_OFF', 
                                 'LL_ON', 'LL_OFF', 
                                 'RL_ON', 'RL_OFF',
                                 'PR', 'OR', 'TR', 'SR',
                                 'ITIS', 'DS', 'MO',
                                 'FSA', 'FSR', 'FSF', 'LSA',
                                 'LSR', 'RSA', 'RSR', 
                                 'SA', 'TO', 'RTS', 'FTS', 'TE', 'ART', '1', '2', '3', '4']

    def run(self):
        loading = ""
        while True:
            if self.start_signal:
                print("Starting " + self.vds_name)
                while self.arduino_connection.readline().decode("utf-8").strip() != "SA":      
                    if self.vds_name == "Training": 
                        self.arduino_connection.write(bytes("train", 'utf-8'))
                    else:
                        self.arduino_connection.write(bytes("test", 'utf-8'))
                    print(loading)
                    loading = loading + "*"
                    self.start_signal = False
                
                self.arduino_started_timestamp = datetime.now()
                self.write_to_spreadsheet(self.trial, "SA", self.arduino_started_timestamp)
                self.signal.emit("SA")
            else:
                if not self.end_signal:
                    decodedStream = self.arduino_connection.readline().decode("utf-8").strip()
                    if decodedStream != '':
                        # need to create a buffer to avoid getting just partial strings
                        self.buffer = self.buffer + decodedStream
                        if self.buffer in self.expected_strings:    
                            self.write_to_spreadsheet(self.trial, self.buffer, datetime.now())
                            if self.buffer == 'TE' or self.buffer == "ART":
                                self.trial = self.trial + 1
                                new_sheet_name = "Trial " + str(self.trial)
                                new_sheet_handler = self.workbook.create_sheet(new_sheet_name)
                                new_sheet_handler.append(self.headers)

                            elif self.buffer == 'RSA':
                                self.right_sensor_activations = self.right_sensor_activations + 1
                            elif self.buffer == 'LSA':
                                self.left_sensor_activations = self.left_sensor_activations + 1
                            elif self.buffer == 'FSA':
                                self.feeder_sensor_activations = self.feeder_sensor_activations + 1

                            self.signal.emit(self.buffer)
                            self.buffer = ""

                        elif self.buffer == "6000" or self.buffer == "12000":
                            self.signal.emit(self.buffer)
                            self.buffer = ""

    def start_trial(self):
        self.start_signal = True

    def end_trial(self, date_end, data, test_detailed_data):
        print(test_detailed_data)
        self.end_signal = True

        # current session detailed information in a new xlsx file
        self.workbook.save(filename=self.arduino_started_timestamp.strftime("%a %d %b %Y %H %M %S ") + str(self.vds_name) + ".xlsx")

        # resumed information added to an existing xlsx file called sessions.xlsx
        try:
            sessions_workbook = load_workbook(filename="sessions.xlsx")
            sheet = sessions_workbook.active

            data_to_sheet = [self.vds_name, self.arduino_started_timestamp, date_end, self.right_sensor_activations, self.left_sensor_activations, self.feeder_sensor_activations, self.trial-1]
            data_to_sheet.extend(data)

            if self.vds_name == "Test":
                for k in test_detailed_data.keys():
                    data_to_sheet.extend(test_detailed_data[k])                               
            sheet.append(data_to_sheet)

            sessions_workbook.save(filename="sessions.xlsx")
        except: # if sessions xlsx was left open have a backup plan
            backup_session_workbook = Workbook()
            sheet = backup_session_workbook.active
            
            data_to_sheet = [self.vds_name, self.arduino_started_timestamp, date_end, self.right_sensor_activations, self.left_sensor_activations, self.feeder_sensor_activations, self.trial-1]
            data_to_sheet.extend(data)

            if self.vds_name == "Test":
                for k in test_detailed_data.keys():
                    data_to_sheet.extend(test_detailed_data[k])                               
            sheet.append(data_to_sheet)

            sessions_workbook.save(filename="backup_sessions" + str(randrange(1000)) + ".xlsx" )


    def write_to_spreadsheet(self, sheet_number, code, timestamp):
        working_sheet = self.workbook.get_sheet_by_name(self.workbook.get_sheet_names()[sheet_number])
        working_sheet.append([code, timestamp.strftime("%H:%M:%S:%f"), str(timestamp - self.arduino_started_timestamp)])